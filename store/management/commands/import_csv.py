import csv
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from store.models import Product, Variant


class Command(BaseCommand):
    help = "Import perfumes from CSV or XLSX into the product catalog"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            default=str(settings.BASE_DIR / "store" / "data" / "JLT_Perfume_List.xlsx"),
            help="CSV/XLSX file path. Defaults to store/data/JLT_Perfume_List.xlsx.",
        )
        parser.add_argument(
            "--reset",
            action="store_true",
            help="Delete existing products and variants before importing.",
        )
        parser.add_argument(
            "--replace-catalog",
            action="store_true",
            help="Delete products that are not present in the source file after updating/creating source products.",
        )
        parser.add_argument(
            "--skip-if-products-exist",
            action="store_true",
            help="Do nothing when at least one product already exists.",
        )

    def _clean_header(self, value):
        return " ".join(str(value or "").strip().upper().split())

    def _clean_cell(self, value):
        if value is None:
            return ""
        return str(value).strip()

    def _read_csv_rows(self, file_path):
        with file_path.open(newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            return [
                {self._clean_header(key): self._clean_cell(value) for key, value in row.items()}
                for row in reader
            ]

    def _read_xlsx_rows(self, file_path):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError(
                "openpyxl is required for .xlsx imports. Run pip install -r requirements.txt."
            ) from exc

        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [self._clean_header(value) for value in next(rows, [])]
        records = []

        for raw_row in rows:
            record = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                record[header] = self._clean_cell(raw_row[index] if index < len(raw_row) else "")
            records.append(record)

        workbook.close()
        return records

    def _read_source_rows(self, file_path):
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._read_csv_rows(file_path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_xlsx_rows(file_path)
        raise CommandError(f"Unsupported import file type: {file_path.suffix}. Use .csv or .xlsx.")

    def _row_value(self, row, *names, default=""):
        for name in names:
            value = row.get(self._clean_header(name))
            if value:
                return value
        return default

    def _normalize_category(self, value):
        normalized = (value or "like").strip().lower()
        aliases = {
            "inspired": "like",
            "inspired fragrances": "like",
            "just like that": "like",
            "jlt": "like",
            "original": "love",
            "original fragrances": "love",
            "just love that": "love",
        }
        return aliases.get(normalized, normalized or "like")

    def _parse_stock(self, value):
        try:
            return max(0, int(float(value)))
        except (TypeError, ValueError):
            return 100

    def _source_key(self, name, brand):
        return (name.strip().casefold(), brand.strip().casefold())

    def handle(self, *args, **options):
        if options["skip_if_products_exist"] and Product.objects.exists():
            self.stdout.write(self.style.WARNING("Products already exist. Skipping product import."))
            return

        file_path = Path(options["path"])
        if not file_path.is_absolute():
            file_path = settings.BASE_DIR / file_path

        if not file_path.exists():
            raise FileNotFoundError(f"Product import file not found: {file_path}")

        rows = self._read_source_rows(file_path)
        source_records = []
        source_keys = set()

        for row in rows:
            name = self._row_value(row, "PRODUCTS", "PRODUCT", "NAME", "INSPIRED BY")
            brand = self._row_value(row, "BRANDS", "BRAND")
            if not name or not brand:
                continue
            key = self._source_key(name, brand)
            if key in source_keys:
                continue
            source_keys.add(key)
            source_records.append((row, name, brand, key))

        if not source_records:
            raise CommandError(f"No valid products found in {file_path}.")

        imported = 0
        updated = 0
        deleted = 0

        with transaction.atomic():
            if options["reset"]:
                Variant.objects.all().delete()
                Product.objects.all().delete()

            for row, name, brand, key in source_records:
                category_value = self._row_value(row, "CATEGORY", "COLLECTION")
                defaults = {
                    "category": self._normalize_category(category_value),
                    "top_note": self._row_value(row, "TOP NOTE", "TOP NOTES"),
                    "middle_note": self._row_value(row, "MIDDLE NOTE", "MIDDLE NOTES", "HEART NOTE", "HEART NOTES"),
                    "base_note": self._row_value(row, "BASE NOTE", "BASE NOTES"),
                    "description": self._row_value(row, "DESCRIPTION", "DESC"),
                    "occasion": self._row_value(row, "OCCASION", "OCCASIONS"),
                    "stock": self._parse_stock(self._row_value(row, "STOCK", "QUANTITY", default="100")),
                }
                product = Product.objects.filter(inspired_by=name, brand=brand).first()
                if product:
                    if not category_value:
                        defaults["category"] = product.category or "like"
                    for field, value in defaults.items():
                        setattr(product, field, value)
                    product.save()
                    created = False
                else:
                    product = Product.objects.create(
                        inspired_by=name,
                        brand=brand,
                        **defaults,
                    )
                    created = True

                if not product.variants.exists():
                    for size, price in Product.DEFAULT_VARIANTS:
                        Variant.objects.create(product=product, size=size, price=price)

                imported += int(created)
                updated += int(not created)

            if options["replace_catalog"] and not options["reset"]:
                stale_ids = []
                for product in Product.objects.only("id", "inspired_by", "brand"):
                    if self._source_key(product.inspired_by, product.brand) not in source_keys:
                        stale_ids.append(product.id)
                if stale_ids:
                    deleted, _ = Product.objects.filter(id__in=stale_ids).delete()

        self.stdout.write(
            self.style.SUCCESS(
                f"Product import complete from {file_path}: "
                f"{imported} created, {updated} updated, {deleted} deleted."
            )
        )
